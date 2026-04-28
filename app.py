"""
app.py — forecast.wbn | Wellbeing Nutrition
v6: 2-month rolling lock, floating calculator, PostgreSQL database layer.

Lock rules:
  In April  → April & May are locked.  Can only refill June.
  In May    → May & June are locked.   Can only refill July.
  Generally → current month AND the immediately following month are locked.
              Earliest refillable month = current_month + 2.
"""
import os, io, smtplib, ssl, threading, datetime, json
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from functools import wraps

from flask import (Flask, render_template, redirect, url_for,
                   session, request, jsonify, send_file)
from dotenv import load_dotenv
from werkzeug.middleware.proxy_fix import ProxyFix
import msal

from excel_handler import (CHANNELS, get_sample_products, save_submission_log)

# ── Optional DB import (graceful fallback to in-memory if not configured) ──
try:
    from db import (
        db_save_submission, db_get_submission, db_get_all_subs_for_quarter,
        db_save_quarter, db_get_quarter, db_revoke_quarter,
        db_save_log, db_get_log,
        db_get_feature_flags, db_set_feature_flag,
        db_get_ticker, db_set_ticker,
        DB_ENABLED,
    )
except ImportError:
    DB_ENABLED = False
    def db_save_submission(*a, **kw): pass
    def db_get_submission(*a, **kw): return None
    def db_get_all_subs_for_quarter(*a, **kw): return {}
    def db_save_quarter(*a, **kw): pass
    def db_get_quarter(*a, **kw): return None
    def db_revoke_quarter(*a, **kw): pass
    def db_save_log(*a, **kw): pass
    def db_get_log(*a, **kw): return []
    def db_get_feature_flags(*a, **kw): return {}
    def db_set_feature_flag(*a, **kw): pass
    def db_get_ticker(*a, **kw): return {"message":"","active":False,"style":"info"}
    def db_set_ticker(*a, **kw): pass

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "forecast-wbn-dev-secret")
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

MAINTENANCE_MODE   = os.getenv("MAINTENANCE_MODE", "false").lower() == "true"
MAINTENANCE_BYPASS = [e.strip().lower() for e in os.getenv("MAINTENANCE_BYPASS_EMAILS","").split(",") if e.strip()]

CLIENT_ID     = os.getenv("CLIENT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")
TENANT_ID     = os.getenv("TENANT_ID")
AUTHORITY     = "https://login.microsoftonline.com/" + (TENANT_ID or "")
REDIRECT_PATH = "/getAToken"
SCOPE         = ["User.Read"]

def get_redirect_uri():
    scheme = "http" if request.host.startswith(("localhost","127.0.0.1")) else "https"
    return url_for("authorized", _external=True, _scheme=scheme)

def _msal_app():
    return msal.ConfidentialClientApplication(
        CLIENT_ID, authority=AUTHORITY, client_credential=CLIENT_SECRET)

def _auth_url():
    return _msal_app().get_authorization_request_url(SCOPE, redirect_uri=get_redirect_uri())

ADMINS           = [e.strip().lower() for e in os.getenv("ADMINS","").split(",") if e.strip()]
FORECAST_MEMBERS = [e.strip().lower() for e in os.getenv("FORECAST_MEMBERS","").split(",") if e.strip()]

def _build_channel_map():
    """
    Returns {email: [channel, ...]}  — supports multiple channels per user.
    CHANNEL_MAP=alice@x.com:D2C,alice@x.com:Amazon,bob@x.com:Retail
    """
    m = {}
    for part in os.getenv("CHANNEL_MAP","").split(","):
        part = part.strip()
        if ":" in part:
            email, ch = part.rsplit(":", 1)
            email = email.strip().lower()
            ch    = ch.strip()
            m.setdefault(email, [])
            if ch not in m[email]:
                m[email].append(ch)
    return m

CHANNEL_MAP = _build_channel_map()

NOTIFY_EMAILS = [e.strip() for e in os.getenv("NOTIFY_EMAILS",
    "nivas@wellbeingnutrition.com,rushikesh.pawar@wellbeingnutrition.com").split(",") if e.strip()]
SMTP_USER = os.getenv("SMTP_USER","")
SMTP_PASS = os.getenv("SMTP_PASS","")

QUARTERS = {
    "Q1": {"label":"Q1 (Apr-Jun)","months":["April","May","June"],           "short":["Apr","May","Jun"]},
    "Q2": {"label":"Q2 (Jul-Sep)","months":["July","August","September"],    "short":["Jul","Aug","Sep"]},
    "Q3": {"label":"Q3 (Oct-Dec)","months":["October","November","December"],"short":["Oct","Nov","Dec"]},
    "Q4": {"label":"Q4 (Jan-Mar)","months":["January","February","March"],   "short":["Jan","Feb","Mar"]},
}

DRR_LABELS = ["7 Days Overall DRR","15 Days Overall DRR","30 Days Overall DRR",
               "45 Days Overall DRR","60 Days Overall DRR"]
DRR_SHORT  = ["7D","15D","30D","45D","60D"]
BASE_COLS_DRR = ["Category","Sub-Category","Product Type","Product Name","SKU","Live/Not Live","Sub Product Type"]

_lock        = threading.Lock()
_quarters    = {}
_submissions = {}
_admin_log   = []
_feature_flags = {
    "load_sample_values": True,
    "download_template":  True,
    "upload_excel_fill":  True,
}
_ticker = {"message": "", "active": False, "style": "info"}

MONTH_NUM = {
    "January":1,"February":2,"March":3,"April":4,
    "May":5,"June":6,"July":7,"August":8,"September":9,
    "October":10,"November":11,"December":12
}

def _get_flags() -> dict:
    """Return current feature flags from DB or in-memory fallback."""
    if DB_ENABLED:
        f = db_get_feature_flags()
        for k, v in _feature_flags.items():
            f.setdefault(k, v)
        return f
    return dict(_feature_flags)

def _set_flag(key: str, value: bool):
    if DB_ENABLED:
        db_set_feature_flag(key, value)
    else:
        with _lock:
            _feature_flags[key] = value

def _get_ticker_data() -> dict:
    if DB_ENABLED:
        return db_get_ticker()
    with _lock:
        return dict(_ticker)

def _set_ticker_data(message: str, active: bool, style: str):
    if DB_ENABLED:
        db_set_ticker(message, active, style)
    else:
        with _lock:
            _ticker.update({"message": message, "active": active, "style": style})

# ── Submission helpers ────────────────────────────────────────────────────────
def _sub_key(qkey, email):
    return qkey + "||" + email.lower()

def _sub_default():
    return dict(
        submitted=False, submitted_at="", submitted_at_dt=None,
        data=None, user_name="", revision=0,
        refill_requested=False, refill_reason="",
        refill_cooldown_until=None,
        excel_bytes=None, file="", channel=""
    )

def _is_channel_keyed(data):
    return isinstance(data, dict) and any(k in CHANNELS for k in data.keys())

def _wrap_legacy_data(raw_data, email, preferred_channel=None):
    if not raw_data or _is_channel_keyed(raw_data):
        return raw_data or {}
    member_channels = CHANNEL_MAP.get((email or '').lower()) or []
    channel = preferred_channel or (member_channels[0] if member_channels else 'Unassigned')
    return {channel: raw_data}

def _sub_get(qkey, email):
    k = _sub_key(qkey, email)
    if DB_ENABLED:
        db_val = db_get_submission(qkey, email)
        if db_val:
            with _lock:
                _submissions[k] = db_val
    with _lock:
        return dict(_submissions.get(k, _sub_default()))

def _sub_set(qkey, email, updates):
    k = _sub_key(qkey, email)
    with _lock:
        if k not in _submissions:
            _submissions[k] = _sub_default()
        _submissions[k].update(updates)
        snap = dict(_submissions[k])
    if DB_ENABLED:
        db_save_submission(qkey, email, snap)

def _sub_reset(qkey, email):
    k = _sub_key(qkey, email)
    default = _sub_default()
    with _lock:
        _submissions[k] = default
    if DB_ENABLED:
        db_save_submission(qkey, email, default)

def _q_get(qkey):
    if DB_ENABLED:
        db_val = db_get_quarter(qkey)
        if db_val:
            with _lock:
                _quarters[qkey] = db_val
    with _lock:
        return dict(_quarters.get(qkey, {}))

def _q_set(qkey, data):
    with _lock:
        _quarters[qkey] = data
    if DB_ENABLED:
        db_save_quarter(qkey, data)

def _q_revoke(qkey):
    with _lock:
        _quarters.pop(qkey, None)
        for k in [k for k in _submissions if k.startswith(qkey+"||")]:
            del _submissions[k]
    if DB_ENABLED:
        db_revoke_quarter(qkey)

def _all_subs_for_quarter(qkey):
    if DB_ENABLED:
        return db_get_all_subs_for_quarter(qkey)
    with _lock:
        result = {}
        for k, v in _submissions.items():
            if k.startswith(qkey+"||"):
                email = k.split("||",1)[1]
                result[email] = dict(v)
        return result

def _log(action, user, detail=""):
    entry = dict(
        timestamp=datetime.datetime.now().strftime("%d %b %Y, %H:%M"),
        action=action, user=user, detail=detail)
    with _lock:
        _admin_log.append(entry)
    if DB_ENABLED:
        db_save_log(entry)

# ── Email ────────────────────────────────────────────────────────────────────
def _send_email(subject, body, attach_name=None, attach_bytes=None):
    if not SMTP_USER or not SMTP_PASS:
        app.logger.warning("SMTP not configured — email skipped.")
        return
    def _work():
        try:
            msg = MIMEMultipart()
            msg["From"] = SMTP_USER
            msg["To"]   = ", ".join(NOTIFY_EMAILS)
            msg["Subject"] = subject
            msg.attach(MIMEText(body, "plain"))
            if attach_name and attach_bytes:
                part = MIMEBase("application","octet-stream")
                part.set_payload(attach_bytes); encoders.encode_base64(part)
                part.add_header("Content-Disposition",'attachment; filename="'+attach_name+'"')
                msg.attach(part)
            ctx = ssl.create_default_context()
            with smtplib.SMTP("smtp.office365.com", 587) as srv:
                srv.ehlo(); srv.starttls(context=ctx); srv.ehlo()
                srv.login(SMTP_USER, SMTP_PASS)
                srv.sendmail(SMTP_USER, NOTIFY_EMAILS, msg.as_string())
        except Exception as e:
            app.logger.error("Email failed: "+str(e))
    threading.Thread(target=_work, daemon=True).start()

# ── Auth decorators ──────────────────────────────────────────────────────────
def _require_login(f):
    @wraps(f)
    def d(*a, **kw):
        if not session.get("user"): return redirect(url_for("login"))
        email = session["user"].get("preferred_username","").lower().strip()
        if MAINTENANCE_MODE and email not in MAINTENANCE_BYPASS:
            return render_template("maintenance.html"), 503
        if email not in ADMINS and email not in FORECAST_MEMBERS:
            return render_template("access_denied.html"), 403
        return f(*a, **kw)
    return d

def _require_admin(f):
    @wraps(f)
    def d(*a, **kw):
        if not session.get("user"): return redirect(url_for("login"))
        if session["user"].get("preferred_username","").lower().strip() not in ADMINS:
            return render_template("access_denied.html"), 403
        return f(*a, **kw)
    return d

def _email():    return session["user"].get("preferred_username","").lower().strip()
def _name():     return session["user"].get("name", _email())
def _is_admin(): return _email() in ADMINS

# ── 2-Month Rolling Lock Logic ────────────────────────────────────────────────
def _get_today():
    """Isolated date function — easy to patch in tests."""
    return datetime.date.today()

def _refill_allowed_months(qkey):
    """
    2-month rolling lock:
      - Current month is LOCKED (cannot refill).
      - Next month is also LOCKED (cannot refill).
      - Only months >= current_month + 2 are available for refill.
    """
    if qkey not in QUARTERS:
        return []
    months = QUARTERS[qkey]["months"]
    today  = _get_today()
    cur_month = today.month
    cur_year  = today.year

    locked_month_nums = set()
    for delta in (0, 1):
        mn = cur_month + delta
        yr = cur_year
        if mn > 12:
            mn -= 12
            yr += 1
        locked_month_nums.add(mn)

    allowed = []
    for m in months:
        mn = MONTH_NUM.get(m, 0)
        if mn == 0:
            continue
        yr = cur_year
        if mn < cur_month and (cur_month - mn) > 6:
            yr = cur_year + 1

        if (yr, mn) > (cur_year, cur_month) and mn not in locked_month_nums:
            allowed.append(m)

    return allowed


def _locked_months_in_quarter(qkey):
    """
    Returns list of months in the quarter that are currently locked
    (current month + next month, but only those that are in the quarter).
    """
    if qkey not in QUARTERS:
        return []
    months = QUARTERS[qkey]["months"]
    today  = _get_today()
    cur_month = today.month
    cur_year  = today.year

    locked = []
    for delta in (0, 1):
        mn = cur_month + delta
        yr = cur_year
        if mn > 12:
            mn -= 12
            yr += 1
        for m in months:
            if MONTH_NUM.get(m, 0) == mn:
                locked.append(m)
    return locked


def _cooldown_active(sub, qkey=None):
    """
    Returns (True, message) if refill is not allowed.
    Uses 2-month rolling lock logic when qkey provided.
    """
    if qkey:
        allowed = _refill_allowed_months(qkey)
        locked  = _locked_months_in_quarter(qkey)
        if not allowed:
            months = QUARTERS.get(qkey, {}).get("months", [])
            if locked:
                return True, ("Months locked: " + ", ".join(locked) +
                              ". No further months available for refill in this quarter.")
            return True, ("All months in this quarter have passed — refill is no longer available.")
        return False, ""
    until_str = sub.get("refill_cooldown_until")
    if not until_str:
        return False, ""
    try:
        until = datetime.date.fromisoformat(until_str)
        today = datetime.date.today()
        if today < until:
            days = (until - today).days
            return True, "Refill locked for " + str(days) + " more day(s)."
    except Exception:
        pass
    return False, ""

# ── Auth routes ───────────────────────────────────────────────────────────────
@app.route("/")
def login():
    if session.get("user"): return redirect(url_for("forecast"))
    return render_template("login.html")

@app.route("/signin")
def signin(): return redirect(_auth_url())

@app.route(REDIRECT_PATH)
def authorized():
    result = _msal_app().acquire_token_by_authorization_code(
        request.args.get("code",""), scopes=SCOPE, redirect_uri=get_redirect_uri())
    if "error" in result:
        return "Login error: "+str(result.get("error_description")), 400
    session["user"] = result.get("id_token_claims")
    email = session["user"].get("preferred_username","").lower().strip()
    if email not in ADMINS and email not in FORECAST_MEMBERS:
        return render_template("access_denied.html"), 403
    return redirect(url_for("forecast"))

@app.route("/logout")
def logout(): session.clear(); return redirect(url_for("login"))

# ── Forecast page ─────────────────────────────────────────────────────────────
@app.route("/forecast")
@_require_login
def forecast():
    email  = _email(); name = _name()
    is_adm = _is_admin()

    q_status = {}
    for qkey in QUARTERS:
        q = _q_get(qkey)
        if not q.get("initiated"):
            q_status[qkey] = {"initiated": False}
        else:
            sub = _sub_get(qkey, email)
            cooldown, cmsg = _cooldown_active(sub, qkey)
            allowed_months = _refill_allowed_months(qkey)
            locked_months  = _locked_months_in_quarter(qkey)
            q_status[qkey] = {
                "initiated":    True,
                "initiated_at": q.get("initiated_at",""),
                "sku_count":    len(q.get("drr_data") or []),
                "submitted":    sub.get("submitted", False),
                "submitted_at": sub.get("submitted_at",""),
                "revision":     sub.get("revision", 0),
                "refill_requested": sub.get("refill_requested", False),
                "cooldown_active":  cooldown,
                "cooldown_msg":     cmsg,
                "allowed_refill_months": allowed_months,
                "locked_months":         locked_months,
                "can_refill":       len(allowed_months) > 0 and sub.get("submitted", False),
            }

    return render_template("forecast.html",
        user_name=name, user_email=email,
        is_admin=is_adm,
        quarters=QUARTERS, drr_labels=DRR_LABELS, drr_short=DRR_SHORT,
        q_status=q_status,
        today=datetime.date.today().strftime("%A, %d %B %Y"),
        db_enabled=DB_ENABLED,
        ticker=_get_ticker_data(),
        flags=_get_flags(),
    )


# ── API: Quarter status poll ──────────────────────────────────────────────────
@app.route("/api/quarter-status")
@_require_login
def api_quarter_status():
    email = _email()
    result = {}
    for qkey in QUARTERS:
        q = _q_get(qkey)
        if not q.get("initiated"):
            result[qkey] = {"initiated": False}
        else:
            sub = _sub_get(qkey, email)
            cooldown, cmsg = _cooldown_active(sub, qkey)
            allowed_months = _refill_allowed_months(qkey)
            locked_months  = _locked_months_in_quarter(qkey)
            result[qkey] = {
                "initiated":    True,
                "initiated_at": q.get("initiated_at",""),
                "sku_count":    len(q.get("drr_data") or []),
                "submitted":    sub.get("submitted", False),
                "submitted_at": sub.get("submitted_at",""),
                "revision":     sub.get("revision", 0),
                "refill_requested": sub.get("refill_requested", False),
                "cooldown_active":  cooldown,
                "cooldown_msg":     cmsg,
                "allowed_refill_months": allowed_months,
                "locked_months":         locked_months,
                "can_refill":       len(allowed_months) > 0 and sub.get("submitted", False),
            }
    return jsonify(result)

# ── API: Load DRR data ─────────────────────────────────────────────────────────
@app.route("/api/load-drr/<qkey>")
@_require_login
def api_load_drr(qkey):
    if qkey not in QUARTERS: return jsonify({"error":"Invalid quarter"}), 400
    q = _q_get(qkey)
    if not q.get("initiated"): return jsonify({"error":"Quarter not yet initiated by admin"}), 404

    email    = _email(); is_adm = _is_admin()
    drr_data = q.get("drr_data") or []
    sub      = _sub_get(qkey, email)
    raw_data = sub.get("data") or {}

    # Channel is now the forecast data scope for both admins and members.
    # Previously admin channel switching only changed DRR reference columns,
    # so one flat forecast value appeared for every channel.
    if is_adm:
        member_channels = CHANNELS[:]
        ch_param = request.args.get("channel") or None
        member_ch = ch_param if ch_param in member_channels else (member_channels[0] if member_channels else None)
    else:
        member_channels = CHANNEL_MAP.get(email) or []
        if not member_channels:
            return jsonify({"error": "You are not assigned to a channel. Contact admin to set CHANNEL_MAP in .env."}), 400
        ch_param = request.args.get("channel") or None
        member_ch = ch_param if ch_param in member_channels else member_channels[0]

    raw_data = _wrap_legacy_data(raw_data, email, sub.get("channel") or member_ch)
    saved = raw_data.get(member_ch, {}) if member_ch else raw_data

    cooldown, cmsg = _cooldown_active(sub, qkey)
    allowed_months = _refill_allowed_months(qkey)
    locked_months  = _locked_months_in_quarter(qkey)

    rows = []
    for row in drr_data:
        r = dict(row)
        drr_all = r.pop("_drr", {})
        row_saved = saved.get(r.get("_row_id",""), {})
        for m in QUARTERS[qkey]["months"]:
            r[m] = row_saved.get(m, "")
        if member_ch:
            ch_drr = drr_all.get(member_ch, {})
            r["_ref"] = {dl: round(ch_drr.get(dl, 0), 2) for dl in DRR_LABELS}
        else:
            r["_refs"] = {}
            for ch in CHANNELS:
                ch_drr = drr_all.get(ch, {})
                r["_refs"][ch] = {dl: round(ch_drr.get(dl, 0), 2) for dl in DRR_LABELS}
        rows.append(r)

    cats    = sorted(set(r.get("Category","")    for r in rows if r.get("Category")))
    subcats = sorted(set(r.get("Sub-Category","") for r in rows if r.get("Sub-Category")))
    ptypes  = sorted(set(r.get("Product Type","") for r in rows if r.get("Product Type")))

    return jsonify({
        "rows":          rows,
        "months":        QUARTERS[qkey]["months"],
        "drr_labels":    DRR_LABELS,
        "drr_short":     DRR_SHORT,
        "user_channel":  member_ch,
        "user_channels": member_channels,
        "all_channels":  None,
        "submitted":     sub.get("submitted", False),
        "submitted_at":  sub.get("submitted_at",""),
        "revision":      sub.get("revision", 0),
        "refill_requested": sub.get("refill_requested", False),
        "cooldown_active":  cooldown,
        "cooldown_msg":     cmsg,
        "allowed_refill_months": allowed_months,
        "locked_months":         locked_months,
        "can_refill":       len(allowed_months) > 0 and sub.get("submitted", False),
        "filter_options": {"categories":cats,"sub_categories":subcats,"product_types":ptypes},
        "flags": _get_flags(),
    })

# ── API: Save draft ────────────────────────────────────────────────────────────
@app.route("/api/save-draft/<qkey>", methods=["POST"])
@_require_login
def api_save_draft(qkey):
    if qkey not in QUARTERS: return jsonify({"error":"Invalid quarter"}), 400
    email = _email()
    sub   = _sub_get(qkey, email)
    if sub["submitted"]:
        return jsonify({"error":"Already submitted — cannot save draft"}), 409

    # FIX: read entire body once so both rows and channel are available
    body = request.json or {}
    rows = body.get("rows", [])
    member_channels = CHANNELS[:] if _is_admin() else (CHANNEL_MAP.get(email) or [])
    ch_param = body.get("channel") or request.args.get("channel") or None
    member_ch = ch_param if (ch_param and ch_param in member_channels) else (member_channels[0] if member_channels else None)

    # Build this channel's data from rows
    data = {row.get("_row_id",""): {m: row.get(m,"") for m in QUARTERS[qkey]["months"]} for row in rows}

    # Merge into channel-keyed store — preserves other channels
    existing_sub = _sub_get(qkey, email)
    existing = _wrap_legacy_data(existing_sub.get("data") or {}, email, existing_sub.get("channel") or member_ch)
    if member_ch:
        existing[member_ch] = data
        merged = existing
    else:
        merged = data  # admin or unassigned — store flat
    _sub_set(qkey, email, {"data": merged, "user_name": _name()})
    return jsonify({"status": "saved", "channel": member_ch})

# ── API: Submit ────────────────────────────────────────────────────────────────
@app.route("/api/submit/<qkey>", methods=["POST"])
@_require_login
def api_submit(qkey):
    if qkey not in QUARTERS: return jsonify({"error":"Invalid quarter"}), 400
    email = _email(); name = _name()
    sub   = _sub_get(qkey, email)

    if sub["submitted"]:
        return jsonify({"error":"Already submitted","submitted_at":sub["submitted_at"]}), 409

    # FIX: read entire body once
    body_data = request.json or {}
    rows = body_data.get("rows", [])
    if not rows: return jsonify({"error":"No data"}), 400

    months = QUARTERS[qkey]["months"]
    errors = []
    for row in rows:
        sku = row.get("SKU") or row.get("Product Name","?")
        for m in months:
            v = row.get(m,"")
            if v == "" or v is None:
                errors.append("["+sku+"] '"+m+"' is empty.")
            else:
                try: float(str(v).replace(",",""))
                except ValueError: errors.append("["+sku+"] '"+m+"' invalid.")
    if errors: return jsonify({"error":"Validation failed","details":errors[:10]}), 422

    data = {row.get("_row_id",""): {m: row.get(m,"") for m in months} for row in rows}

    # Multi-channel support: each channel stored separately
    member_channels = CHANNELS[:] if _is_admin() else (CHANNEL_MAP.get(email) or [])
    ch_param = body_data.get("channel") or request.args.get("channel") or None
    member_ch = ch_param if (ch_param and ch_param in member_channels) else (member_channels[0] if member_channels else None)

    # Merge into channel-keyed store — Channel B submit does NOT erase Channel A
    existing_sub = _sub_get(qkey, email)
    existing_data = _wrap_legacy_data(existing_sub.get("data") or {}, email, existing_sub.get("channel") or member_ch)
    if member_ch:
        existing_data[member_ch] = data
        merged_data = existing_data
    else:
        merged_data = data  # admin / no-channel

    q_master  = _q_get(qkey)
    drr_lookup = {p["_row_id"]: p.get("_drr", {}) for p in (q_master.get("drr_data") or [])}
    for row in rows:
        rid     = row.get("_row_id", "")
        drr_all = drr_lookup.get(rid, {})
        if drr_all:
            row["_refs"] = {
                ch: {dl: round(drr_all.get(ch, {}).get(dl, 0), 2) for dl in DRR_LABELS}
                for ch in CHANNELS
            }
            if member_ch:
                row["_ref"] = row["_refs"].get(member_ch, {})

    buf = io.BytesIO()
    # If the user has forecast data for more than one channel, build an all-channels
    # Excel so the downloaded file contains the full picture across every channel.
    active_chs_in_data = [ch for ch in CHANNELS if ch in merged_data and merged_data[ch]]
    if len(active_chs_in_data) > 1:
        _save_submission_excel_multi_channel(
            merged_data, q_master.get("drr_data") or [],
            name, qkey, months, buf)
    else:
        _save_submission_excel(rows, name, qkey, months, buf, member_channel=member_ch)
    buf.seek(0)
    eb   = buf.read()

    now  = datetime.datetime.now()
    at   = now.strftime("%d %b %Y, %H:%M")
    rev  = sub["revision"] + 1
    fn   = ("Forecast_"+qkey+"_"+datetime.date.today().strftime("%Y_%m_%d")
            +"_"+name.replace(" ","_")+((("_r"+str(rev)) if rev > 1 else ""))+".xlsx")

    allowed_refill_months = _refill_allowed_months(qkey)
    locked_months         = _locked_months_in_quarter(qkey)
    can_refill = len(allowed_refill_months) > 0

    _sub_set(qkey, email, dict(
        submitted=True, submitted_at=at, submitted_at_dt=now.isoformat(),
        data=merged_data, user_name=name, revision=rev, channel=member_ch,
        refill_requested=False, refill_reason="",
        refill_cooldown_until=None,
        excel_bytes=eb, file=fn
    ))

    try: save_submission_log(name, fn, len(rows), ".")
    except: pass
    _log("Submission", name, qkey+" Rev "+str(rev)+" — "+str(len(rows))+" rows")

    lock_note = ""
    if locked_months:
        lock_note = "Locked months (cannot refill): " + ", ".join(locked_months) + "\n"
    refill_note = ("Refill available for: "+", ".join(allowed_refill_months)) if can_refill else "No refill available (all months locked or passed)"

    _send_email(
        subject="Forecast Submitted — "+name+" / "+qkey+" — "+datetime.date.today().strftime("%d %b %Y"),
        body=("Submitted by : "+name+" ("+email+")\n"
              "Quarter      : "+qkey+" — "+QUARTERS[qkey]["label"]+"\n"
              "Date & Time  : "+at+"\n"
              "Products     : "+str(len(rows))+"\n"
              "Revision     : "+str(rev)+"\n"
              +lock_note+refill_note+"\n\n— forecast.wbn | Wellbeing Nutrition"),
        attach_name=fn, attach_bytes=eb)

    return jsonify({"status":"success","submitted_at":at,"filename":fn,"revision":rev,
                    "allowed_refill_months":allowed_refill_months,
                    "locked_months":locked_months,
                    "can_refill":can_refill})

# ── API: Download own submission ───────────────────────────────────────────────
@app.route("/api/download-submission/<qkey>")
@_require_login
def api_download_submission(qkey):
    email = _email()
    sub   = _sub_get(qkey, email)
    if not sub.get("submitted"): return jsonify({"error":"No submission"}), 404
    eb = sub.get("excel_bytes")
    if not eb: return jsonify({"error":"File unavailable"}), 404
    buf = io.BytesIO(eb); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=sub.get("file","submission.xlsx"),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── API: Refill request ────────────────────────────────────────────────────────
@app.route("/api/request-refill/<qkey>", methods=["POST"])
@_require_login
def api_request_refill(qkey):
    email = _email(); name = _name()
    sub   = _sub_get(qkey, email)

    if not sub.get("submitted"):
        return jsonify({"error":"Nothing submitted yet"}), 400

    cooldown, msg = _cooldown_active(sub, qkey)
    if cooldown:
        return jsonify({"error": "Refill not available", "reason": msg}), 403

    allowed_months = _refill_allowed_months(qkey)
    if not allowed_months:
        locked = _locked_months_in_quarter(qkey)
        reason = "All months are locked ("+", ".join(locked)+") or have passed." if locked else "No months left to refill."
        return jsonify({"error": "Refill not available", "reason": reason}), 403

    if sub.get("refill_requested"):
        return jsonify({"error":"Refill already requested — awaiting admin approval"}), 409

    reason = (request.json or {}).get("reason","").strip()
    if not reason: return jsonify({"error":"Please provide a reason"}), 400

    _sub_set(qkey, email, {"refill_requested":True, "refill_reason":reason})
    _log("Refill Request", name, qkey+" — "+reason)
    _send_email(
        subject="Refill Request — "+name+" / "+qkey,
        body=(name+" ("+email+") has requested a forecast refill.\n\n"
              "Quarter        : "+qkey+" — "+QUARTERS[qkey]["label"]+"\n"
              "Reason         : "+reason+"\n"
              "Refillable for : "+", ".join(allowed_months)+"\n\n"
              "Log in to the Admin Panel to approve or deny.\n\n— forecast.wbn"))
    return jsonify({"status":"requested", "allowed_months": allowed_months})

# ── API: Download template ─────────────────────────────────────────────────────
@app.route("/api/download-template/<qkey>")
@_require_login
def api_download_template(qkey):
    if qkey not in QUARTERS: return jsonify({"error":"Invalid quarter"}), 400
    buf = io.BytesIO(); _create_quarter_template(qkey, buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="Forecast_Template_"+qkey+".xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── Admin panel ────────────────────────────────────────────────────────────────
@app.route("/admin")
@_require_admin
def admin_panel():
    if DB_ENABLED:
        log = list(reversed(db_get_log(50)))
    else:
        with _lock: log = list(reversed(_admin_log[-50:]))

    q_info = {}
    for qkey, qmeta in QUARTERS.items():
        q    = _q_get(qkey)
        subs = _all_subs_for_quarter(qkey) if q.get("initiated") else {}
        all_members = list(set(ADMINS + FORECAST_MEMBERS))
        submitted_count  = sum(1 for s in subs.values() if s.get("submitted"))
        pending_refills  = sum(1 for s in subs.values() if s.get("refill_requested"))
        member_rows = []
        for m_email in sorted(all_members):
            if m_email in ADMINS: continue
            s = subs.get(m_email, _sub_default())
            cooldown, cmsg = _cooldown_active(s, qkey)
            am = _refill_allowed_months(qkey)
            lm = _locked_months_in_quarter(qkey)
            member_rows.append({
                "email":        m_email,
                "user_name":    s.get("user_name", m_email.split("@")[0]),
                "submitted":    s.get("submitted", False),
                "submitted_at": s.get("submitted_at",""),
                "revision":     s.get("revision", 0),
                "refill_requested":      s.get("refill_requested", False),
                "refill_reason":         s.get("refill_reason",""),
                "cooldown_active":       cooldown,
                "cooldown_msg":          cmsg,
                "allowed_refill_months": am,
                "locked_months":         lm,
                "can_refill":            len(am) > 0 and s.get("submitted", False),
            })
        q_info[qkey] = {
            "label":           qmeta["label"],
            "initiated":       q.get("initiated", False),
            "initiated_at":    q.get("initiated_at",""),
            "sku_count":       len(q.get("drr_data") or []),
            "submitted_count": submitted_count,
            "pending_refills": pending_refills,
            "total_members":   len(member_rows),
            "members":         member_rows,
        }
    total_submitted = sum(q["submitted_count"] for q in q_info.values())
    pending_refills = sum(q["pending_refills"] for q in q_info.values())
    return render_template("admin.html",
        user_name=_name(), today=datetime.date.today().strftime("%A, %d %B %Y"),
        quarters=QUARTERS, q_info=q_info,
        total_submitted=total_submitted, pending_refills=pending_refills,
        activity_log=log, db_enabled=DB_ENABLED,
        ticker=_get_ticker_data(),
        flags=_get_flags(),
    )

@app.route("/admin/api/initiate-quarter", methods=["POST"])
@_require_admin
def admin_initiate_quarter():
    qkey = request.form.get("quarter","")
    if qkey not in QUARTERS: return jsonify({"error":"Invalid quarter"}), 400
    if "file" not in request.files: return jsonify({"error":"No file"}), 400
    f = request.files["file"]
    if not f.filename.endswith((".xlsx",".xls")): return jsonify({"error":"Only .xlsx/.xls"}), 400
    try: drr_data, channels_found = _parse_drr_excel(io.BytesIO(f.read()))
    except Exception as e: return jsonify({"error":str(e)}), 400
    _q_set(qkey, {"initiated":True, "initiated_at":datetime.date.today().strftime("%d %b %Y"),
                   "drr_data":drr_data, "channels_found":channels_found})
    _log("Quarter Initiated", _name(), qkey+" — "+str(len(drr_data))+" SKUs — visible to all members")
    return jsonify({"status":"ok","sku_count":len(drr_data),"channels_found":channels_found})

@app.route("/admin/api/revoke-quarter", methods=["POST"])
@_require_admin
def admin_revoke_quarter():
    qkey = (request.json or {}).get("quarter","")
    if qkey not in QUARTERS: return jsonify({"error":"Invalid"}), 400
    _q_revoke(qkey); _log("Quarter Revoked", _name(), qkey)
    return jsonify({"status":"ok"})

@app.route("/admin/api/approve-refill", methods=["POST"])
@_require_admin
def admin_approve_refill():
    data  = request.json or {}
    qkey  = data.get("quarter","")
    email = data.get("email","").lower().strip()
    sub   = _sub_get(qkey, email)
    if not sub.get("refill_requested"):
        return jsonify({"error":"No pending request"}), 400
    _sub_set(qkey, email, {
        "submitted":       False,
        "refill_requested":False,
        "refill_reason":   "",
        "refill_cooldown_until": None,
    })
    _log("Refill Approved", _name(), qkey+" for "+email)
    return jsonify({"status":"approved"})

@app.route("/admin/api/deny-refill", methods=["POST"])
@_require_admin
def admin_deny_refill():
    data  = request.json or {}
    qkey  = data.get("quarter","")
    email = data.get("email","").lower().strip()
    _sub_set(qkey, email, {"refill_requested":False,"refill_reason":""})
    _log("Refill Denied", _name(), qkey+" for "+email)
    return jsonify({"status":"denied"})

@app.route("/admin/api/force-unlock", methods=["POST"])
@_require_admin
def admin_force_unlock():
    data  = request.json or {}
    qkey  = data.get("quarter","")
    email = data.get("email","").lower().strip()
    _sub_set(qkey, email, {
        "submitted":False, "refill_requested":False,
        "refill_cooldown_until":None,
    })
    _log("Force Unlock", _name(), qkey+" for "+email)
    return jsonify({"status":"unlocked"})

@app.route("/admin/api/export-quarter/<qkey>")
@_require_admin
def admin_export_quarter(qkey):
    if qkey not in QUARTERS: return jsonify({"error":"Invalid"}), 400
    q = _q_get(qkey)
    if not q.get("initiated"): return jsonify({"error":"Not initiated"}), 404
    buf = io.BytesIO(); _export_quarter_excel(qkey, q, buf); buf.seek(0)
    fn = "DRR_Forecast_"+qkey+"_"+datetime.date.today().strftime("%Y-%m-%d")+".xlsx"
    return send_file(buf, as_attachment=True, download_name=fn,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/admin/api/download-member-submission/<qkey>/<path:email>")
@_require_admin
def admin_download_member(qkey, email):
    sub = _sub_get(qkey, email.lower())
    eb  = sub.get("excel_bytes")
    if not eb: return jsonify({"error":"No file"}), 404
    buf = io.BytesIO(eb); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=sub.get("file","submission.xlsx"),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/admin/api/set-ticker", methods=["POST"])
@_require_admin
def admin_set_ticker():
    data    = request.json or {}
    message = data.get("message", "").strip()
    active  = bool(data.get("active", False))
    style   = data.get("style", "info")
    if style not in ("info", "warn", "success", "danger"):
        style = "info"
    _set_ticker_data(message, active, style)
    _log("Ticker Updated", _name(),
         ("Active: " if active else "Deactivated. ") + message[:60])
    return jsonify({"status": "ok"})


@app.route("/admin/api/set-feature-flag", methods=["POST"])
@_require_admin
def admin_set_feature_flag():
    data    = request.json or {}
    flag    = data.get("flag", "")
    enabled = bool(data.get("enabled", True))
    VALID   = {"load_sample_values", "download_template", "upload_excel_fill"}
    if flag not in VALID:
        return jsonify({"error": "Unknown flag"}), 400
    _set_flag(flag, enabled)
    _log("Feature Flag", _name(), f"{flag} → {'ON' if enabled else 'OFF'}")
    return jsonify({"status": "ok", "flag": flag, "enabled": enabled})


@app.route("/api/get-ticker")
@_require_login
def api_get_ticker():
    """Polled by forecast page every 60 s to pick up live ticker changes."""
    return jsonify(_get_ticker_data())


# ── API: Upload Excel fill ─────────────────────────────────────────────────────
@app.route("/api/upload-excel-fill/<qkey>", methods=["POST"])
@_require_login
def api_upload_excel_fill(qkey):
    """
    Upload a filled Excel file and merge forecast values into the member's draft.
    Channel enforcement: only reads columns for the member's active channel.
    """
    flags = _get_flags()
    if not flags.get("upload_excel_fill", True):
        return jsonify({"error": "Excel upload is currently disabled by admin."}), 403

    if qkey not in QUARTERS:
        return jsonify({"error": "Invalid quarter"}), 400

    q = _q_get(qkey)
    if not q.get("initiated"):
        return jsonify({"error": "Quarter not initiated"}), 404

    email  = _email()
    is_adm = _is_admin()
    sub    = _sub_get(qkey, email)

    if sub.get("submitted"):
        return jsonify({"error": "Already submitted — cannot upload."}), 409

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded."}), 400

    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Only .xlsx / .xls files are accepted."}), 400

    months   = QUARTERS[qkey]["months"]
    drr_data = q.get("drr_data") or []
    saved    = sub.get("data") or {}

    # Determine channel scope
    if is_adm:
        ch_param_upload = request.args.get("channel") or None
        member_ch = ch_param_upload if ch_param_upload in CHANNELS else (CHANNELS[0] if CHANNELS else None)
    else:
        member_channels_upload = CHANNEL_MAP.get(email) or []
        if not member_channels_upload:
            return jsonify({
                "error": "You are not assigned to a channel. "
                         "Ask admin to set your channel in CHANNEL_MAP."
            }), 400
        ch_param_up = request.args.get("channel") or None
        member_ch = ch_param_up if ch_param_up in member_channels_upload else member_channels_upload[0]

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active
        all_xl_rows = list(ws.iter_rows(values_only=True))

        if len(all_xl_rows) < 2:
            return jsonify({"error": "File is empty or too small."}), 422

        header_row_idx = None
        month_col_map  = {}
        _best_score    = 0

        for ri, row in enumerate(all_xl_rows):
            cells = [str(c or "").strip() for c in row]

            if member_ch:
                ch_map = {}
                for ci, cell in enumerate(cells):
                    for m in months:
                        if cell.lower() == f"{member_ch} {m}".lower():
                            ch_map[m] = ci
                if len(ch_map) > _best_score:
                    _best_score    = len(ch_map)
                    header_row_idx = ri
                    month_col_map  = ch_map
                    if len(ch_map) == len(months):
                        break

            plain_map = {}
            for ci, cell in enumerate(cells):
                for m in months:
                    if cell.strip() == m:
                        plain_map[m] = ci
            if len(plain_map) == len(months) and len(plain_map) > _best_score:
                _best_score    = len(plain_map)
                header_row_idx = ri
                month_col_map  = plain_map
                if not member_ch:
                    break

        if not month_col_map:
            if member_ch:
                prefixed = ", ".join(f"{member_ch} {m}" for m in months)
                plain    = ", ".join(months)
                return jsonify({
                    "error": (
                        f"No columns found for your channel '{member_ch}'. "
                        f"The file must contain either:\n"
                        f"  • Plain month names: {plain}\n"
                        f"  • Channel-prefixed names: {prefixed}\n"
                        f"Please download the template from the forecast page "
                        f"and fill in the correct columns."
                    )
                }), 422
            else:
                return jsonify({
                    "error": (
                        f"Could not find month columns ({', '.join(months)}) in the file. "
                        "Please download the template and fill that in."
                    )
                }), 422

        missing_months = [m for m in months if m not in month_col_map]
        if missing_months:
            if member_ch:
                missing_cols = " / ".join(
                    f"'{member_ch} {m}' or '{m}'" for m in missing_months
                )
                return jsonify({
                    "error": (
                        f"Columns missing for channel '{member_ch}': {missing_cols}. "
                        f"Please make sure all {len(months)} month columns are present."
                    )
                }), 422
            else:
                return jsonify({
                    "error": f"Some month columns are missing: {', '.join(missing_months)}."
                }), 422

        sku_col = None
        if header_row_idx is not None:
            hdr = [str(c or "").strip() for c in all_xl_rows[header_row_idx]]
            for ci, cell in enumerate(hdr):
                if cell.upper() in ("SKU", "PRODUCT NAME / SKU", "PRODUCT NAME", "PRODUCT"):
                    sku_col = ci
                    break

        sku_to_rowid  = {}
        name_to_rowid = {}
        for prod in drr_data:
            sku  = (prod.get("SKU")          or "").strip()
            name = (prod.get("Product Name") or "").strip()
            rid  = prod["_row_id"]
            if sku:  sku_to_rowid[sku.lower()]   = rid
            if name: name_to_rowid[name.lower()]  = rid

        SKIP_PATTERNS = {"instruction", "do not edit", "(example)", "fill in",
                         "instructions:", "note:"}

        filled         = 0
        zeroed         = 0
        skipped        = 0
        matched_rowids = set()

        merged = dict(saved)

        for row in all_xl_rows[header_row_idx + 1:]:
            cells = [str(c or "").strip() for c in row]

            if not any(cells):
                continue

            first_cell = cells[0].lower() if cells else ""
            if any(p in first_cell for p in SKIP_PATTERNS):
                continue

            row_id = None

            if sku_col is not None and sku_col < len(cells):
                sku_val = cells[sku_col].lower()
                row_id = sku_to_rowid.get(sku_val) or name_to_rowid.get(sku_val)
                if not row_id and sku_val:
                    for k, v in sku_to_rowid.items():
                        if k and sku_val and (k in sku_val or sku_val in k):
                            row_id = v
                            break

            if not row_id:
                for cell_val in cells:
                    cv = cell_val.lower()
                    row_id = (sku_to_rowid.get(cv) or name_to_rowid.get(cv))
                    if row_id:
                        break

            if not row_id:
                skipped += 1
                continue

            if row_id not in merged:
                merged[row_id] = {}

            matched_rowids.add(row_id)

            for month, col_idx in month_col_map.items():
                if col_idx < len(row):
                    raw     = row[col_idx]
                    raw_str = str(raw).strip() if raw is not None else ""

                    if raw_str in ("", "-", "—", "None"):
                        merged[row_id][month] = 0
                        zeroed += 1
                    else:
                        try:
                            val = float(raw_str.replace(",", ""))
                            merged[row_id][month] = val
                            if val != 0:
                                filled += 1
                            else:
                                zeroed += 1
                        except ValueError:
                            merged[row_id][month] = 0
                            zeroed += 1
                else:
                    merged[row_id][month] = 0
                    zeroed += 1

        if not matched_rowids:
            ch_hint = f" for channel '{member_ch}'" if member_ch else ""
            return jsonify({
                "error": (
                    f"No matching products were found{ch_hint} in the uploaded file. "
                    f"Make sure SKU or Product Name values match the master list, "
                    f"and that the month columns ({', '.join(months)}) are present."
                )
            }), 422

        # FIX: Store under channel key — preserves data for other channels
        existing_sub = _sub_get(qkey, email)
        existing_upload = _wrap_legacy_data(existing_sub.get("data") or {}, email, existing_sub.get("channel") or member_ch)
        member_channels_all = CHANNELS[:] if is_adm else (CHANNEL_MAP.get(email) or [])
        if member_ch:
            existing_upload[member_ch] = merged
            final_merged = existing_upload
        else:
            final_merged = merged
        _sub_set(qkey, email, {"data": final_merged, "user_name": _name()})

        ch_label      = f" [{member_ch}]" if member_ch else ""
        products_hit  = len(matched_rowids)
        zero_note     = f", {zeroed} cell{'s' if zeroed != 1 else ''} set to 0" if zeroed else ""

        return jsonify({
            "status":  "merged",
            "filled":  filled,
            "zeroed":  zeroed,
            "skipped": skipped,
            "channel": member_ch,
            "message": (
                f"Imported {products_hit} product{'s' if products_hit != 1 else ''}{ch_label} "
                f"({filled} non-zero value{'s' if filled != 1 else ''}{zero_note}). "
                + (f"{skipped} row{'s' if skipped != 1 else ''} skipped "
                   f"(SKU not found in master list). " if skipped else "")
                + "Review and submit when ready."
            ),
        })

    except Exception as e:
        return jsonify({"error": f"Could not parse file: {str(e)}"}), 422


# ── Excel helpers ──────────────────────────────────────────────────────────────
def _parse_drr_excel(source):
    import openpyxl
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 3:
        raise ValueError("Sheet too small — need 2 header rows + data.")
    h0 = [str(c or "").strip() for c in all_rows[0]]
    h1 = [str(c or "").strip() for c in all_rows[1]]
    base_idx = {}
    for col in BASE_COLS_DRR:
        try: base_idx[col] = h1.index(col)
        except ValueError: pass
    col_map = {}; cur_ch = None
    for c, cv in enumerate(h0):
        if cv: cur_ch = cv
        lbl = h1[c] if c < len(h1) else ""
        if lbl in DRR_LABELS and cur_ch:
            col_map[c] = {"ch":cur_ch,"label":lbl}
    rows = []
    for ri, raw in enumerate(all_rows[2:], start=2):
        row_str = [str(c or "").strip() for c in raw]
        sku  = row_str[base_idx["SKU"]] if "SKU" in base_idx and base_idx["SKU"]<len(row_str) else ""
        name = row_str[base_idx["Product Name"]] if "Product Name" in base_idx and base_idx["Product Name"]<len(row_str) else ""
        if not sku and not name: continue
        obj = {"_row_id":"r"+str(ri)}
        for col in BASE_COLS_DRR:
            obj[col] = row_str[base_idx[col]] if col in base_idx and base_idx[col]<len(row_str) else ""
        obj["_drr"] = {}
        for c, info in col_map.items():
            ch_name = info["ch"]
            if ch_name not in obj["_drr"]: obj["_drr"][ch_name] = {}
            try:
                val = raw[c] if c < len(raw) else None
                obj["_drr"][ch_name][info["label"]] = round(float(val or 0),2)
            except: obj["_drr"][ch_name][info["label"]] = 0.0
        rows.append(obj)
    channels_found = list({info["ch"] for info in col_map.values()})
    return rows, channels_found

def _create_quarter_template(qkey, dest):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    months   = QUARTERS[qkey]["months"]
    bdr_hdr  = Border(**{s: Side(style="thin", color="FFFFFF") for s in ["left","right","top","bottom"]})
    bdr_data = Border(**{s: Side(style="thin", color="D1D5DB") for s in ["left","right","top","bottom"]})
    bdr_mon  = Border(**{s: Side(style="thin", color="6EE7B7") for s in ["left","right","top","bottom"]})

    PROD_COLS = ["Category", "Sub-Category", "Product Type", "Product Name",
                 "SKU", "Live/Not Live", "Sub Product Type"]
    all_headers = PROD_COLS + months

    wb = Workbook()
    ws = wb.active
    ws.title = "Forecast " + qkey

    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(PROD_COLS))
    hdr_prod = ws.cell(row=1, column=1, value="PRODUCT INFO (do not edit)")
    hdr_prod.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    hdr_prod.fill      = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
    hdr_prod.alignment = Alignment(horizontal="center", vertical="center")
    hdr_prod.border    = bdr_hdr

    ws.merge_cells(start_row=1, start_column=len(PROD_COLS)+1,
                   end_row=1, end_column=len(all_headers))
    hdr_mon = ws.cell(row=1, column=len(PROD_COLS)+1,
                      value="FORECAST — " + QUARTERS[qkey]["label"] + "  ← fill these in")
    hdr_mon.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    hdr_mon.fill      = PatternFill(start_color="40916C", end_color="40916C", fill_type="solid")
    hdr_mon.alignment = Alignment(horizontal="center", vertical="center")
    hdr_mon.border    = bdr_hdr
    ws.row_dimensions[1].height = 22

    for ci, h in enumerate(all_headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        is_month = h in months
        cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill      = PatternFill(
            start_color="40916C" if is_month else "2D6A4F",
            end_color  ="40916C" if is_month else "2D6A4F",
            fill_type  ="solid"
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = bdr_hdr
    ws.row_dimensions[2].height = 28

    q        = _q_get(qkey)
    drr_data = q.get("drr_data") or []

    if drr_data:
        data_rows = drr_data
    else:
        data_rows = [
            {"Category":"(Example)","Sub-Category":"","Product Type":"",
             "Product Name":"Product A","SKU":"SKU-001",
             "Live/Not Live":"Live","Sub Product Type":""},
            {"Category":"(Example)","Sub-Category":"","Product Type":"",
             "Product Name":"Product B","SKU":"SKU-002",
             "Live/Not Live":"Live","Sub Product Type":""},
            {"Category":"(Example)","Sub-Category":"","Product Type":"",
             "Product Name":"Product C","SKU":"SKU-003",
             "Live/Not Live":"Live","Sub Product Type":""},
        ]

    for ri, prod in enumerate(data_rows, 3):
        even = ri % 2 == 0
        for ci, col in enumerate(all_headers, 1):
            is_month = col in months
            if is_month:
                val    = ""
                bg     = "F0FFF4" if even else "DCFCE7"
                cell   = ws.cell(row=ri, column=ci, value=val)
                cell.border = bdr_mon
                cell.number_format = "#,##0.00"
            else:
                val  = prod.get(col, "")
                bg   = "F8FAF9" if even else "FFFFFF"
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = bdr_data
            cell.fill      = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            cell.alignment = Alignment(horizontal="center" if is_month else "left",
                                       vertical="center")
            cell.font      = Font(name="Calibri", size=9,
                                  bold=(col == "Product Name"))
        ws.row_dimensions[ri].height = 18

    widths = {"Category":16,"Sub-Category":16,"Product Type":16,
              "Product Name":28,"SKU":16,"Live/Not Live":12,"Sub Product Type":14}
    for ci, col in enumerate(all_headers, 1):
        ltr = get_column_letter(ci)
        ws.column_dimensions[ltr].width = widths.get(col, 14)

    ws.freeze_panes = get_column_letter(len(PROD_COLS)+1) + "3"

    note_row = len(data_rows) + 4
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=len(all_headers))
    note = ws.cell(row=note_row, column=1,
                   value=("INSTRUCTIONS: Fill in the " + ", ".join(months) +
                          " columns (highlighted green) with your forecast quantities. "
                          "Do not edit the product info columns. "
                          "Save and upload via the 'Upload Excel' button on the forecast page."))
    note.font      = Font(name="Calibri", size=9, italic=True, color="6B7280")
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[note_row].height = 32

    wb.save(dest)


def _save_submission_excel(rows, username, qkey, months, dest, member_channel=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    bdr = Border(**{s: Side(style="thin", color="DEE2E6")
                    for s in ["left","right","top","bottom"]})

    wb = Workbook()
    ws = wb.active
    ws.title = "Forecast Submission"

    sample_row    = rows[0] if rows else {}
    has_single_ch = "_ref"  in sample_row
    has_multi_ch  = "_refs" in sample_row
    drr_channels  = list(sample_row.get("_refs", {}).keys()) if has_multi_ch else []

    if has_multi_ch and drr_channels:
        drr_header_groups = drr_channels
        n_drr_cols = len(drr_channels) * len(DRR_SHORT)
    elif has_single_ch:
        ch_label = member_channel or "Channel"
        drr_header_groups = [ch_label]
        n_drr_cols = len(DRR_SHORT)
    else:
        drr_header_groups = []
        n_drr_cols = 0

    total_cols = len(BASE_COLS_DRR) + n_drr_cols + len(months)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    tc = ws.cell(row=1, column=1,
                 value="Forecast — " + qkey + " — " + username +
                       " — " + datetime.date.today().strftime("%d %b %Y"))
    tc.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
    tc.fill      = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    col = 1
    def _mhdr(r, c, span, text, bg, txt_color="FFFFFF"):
        if span > 1:
            ws.merge_cells(start_row=r, start_column=c,
                           end_row=r, end_column=c+span-1)
        cell = ws.cell(row=r, column=c, value=text)
        cell.font      = Font(bold=True, color=txt_color, name="Calibri", size=9)
        cell.fill      = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = bdr

    _mhdr(2, col, len(BASE_COLS_DRR), "PRODUCT INFO", "1B4332")
    col += len(BASE_COLS_DRR)

    if drr_header_groups:
        for ch in drr_header_groups:
            _mhdr(2, col, len(DRR_SHORT), "DRR — " + ch, "3730A3")
            col += len(DRR_SHORT)

    _mhdr(2, col, len(months), "FORECAST — " + QUARTERS[qkey]["label"], "40916C")
    ws.row_dimensions[2].height = 22

    col = 1
    for h in BASE_COLS_DRR:
        cell = ws.cell(row=3, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
        cell.fill      = PatternFill(start_color="2D6A4F", end_color="2D6A4F", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = bdr
        col += 1

    if drr_header_groups:
        for ch in drr_header_groups:
            for short_lbl in DRR_SHORT:
                cell = ws.cell(row=3, column=col, value=short_lbl)
                cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
                cell.fill      = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = bdr
                col += 1

    for m in months:
        cell = ws.cell(row=3, column=col, value=m)
        cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
        cell.fill      = PatternFill(start_color="40916C", end_color="40916C", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = bdr
        col += 1
    ws.row_dimensions[3].height = 26

    for ri, row in enumerate(rows, 4):
        even = ri % 2 == 0
        col = 1
        for h in BASE_COLS_DRR:
            cell = ws.cell(row=ri, column=col, value=row.get(h, ""))
            cell.fill      = PatternFill(start_color="D8F3DC" if even else "E8F5EC",
                                         end_color="D8F3DC" if even else "E8F5EC",
                                         fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border    = bdr
            cell.font      = Font(name="Calibri", size=9)
            col += 1

        if drr_header_groups:
            for ch in drr_header_groups:
                if has_multi_ch:
                    refs = row.get("_refs", {}).get(ch, {})
                else:
                    refs = row.get("_ref", {})
                for full_lbl in DRR_LABELS:
                    v    = refs.get(full_lbl, "")
                    cell = ws.cell(row=ri, column=col,
                                   value=round(float(v), 2) if v else "")
                    cell.fill      = PatternFill(start_color="E0E7FF" if even else "EEF2FF",
                                                 end_color="E0E7FF" if even else "EEF2FF",
                                                 fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border    = bdr
                    cell.font      = Font(name="Calibri", size=9)
                    col += 1

        for m in months:
            v    = row.get(m, "")
            cell = ws.cell(row=ri, column=col, value=v if v != "" else "")
            cell.fill      = PatternFill(start_color="DCFCE7" if even else "F0FFF4",
                                         end_color="DCFCE7" if even else "F0FFF4",
                                         fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = bdr
            cell.font      = Font(name="Calibri", size=9)
            if v != "":
                cell.number_format = "#,##0.00"
            col += 1

        ws.row_dimensions[ri].height = 18

    for ci in range(1, len(BASE_COLS_DRR)+1):
        ws.column_dimensions[get_column_letter(ci)].width = 16
    drr_start = len(BASE_COLS_DRR)+1
    for ci in range(drr_start, drr_start + n_drr_cols):
        ws.column_dimensions[get_column_letter(ci)].width = 8
    for ci in range(drr_start + n_drr_cols, total_cols+1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    ws.freeze_panes = get_column_letter(len(BASE_COLS_DRR)+1) + "4"
    wb.save(dest)


def _save_submission_excel_multi_channel(merged_data, drr_data, username, qkey, months, dest):
    """
    Build a submission Excel that shows ALL channels the user has submitted.
    merged_data : { channel_name -> { row_id -> { month -> value } } }
    drr_data    : list of product dicts (from q_master["drr_data"])
    Each channel gets its own group of forecast-month columns, labelled with
    the channel name as the group header.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Colours
    C_DARK   = "1B4332"
    C_MID    = "2D6A4F"
    C_PALE   = "D8F3DC"
    C_LIGHT  = "E8F5EC"
    C_WHITE  = "FFFFFF"
    C_GREY   = "F0FFF4"

    # Channel palette — cycle through distinct greens/teals for each channel
    CHANNEL_PALETTES = [
        ("40916C", "D8F3DC", "E8F5EC"),   # green
        ("1D6A96", "D0EAF8", "E3F3FC"),   # blue
        ("7B3FA0", "EAD5F5", "F3E8FC"),   # purple
        ("C05621", "FDEBD0", "FEF3E4"),   # orange
        ("B7451F", "FDDBD0", "FEF0EB"),   # red-orange
        ("1A7A5E", "C8F0E4", "DFF7F0"),   # teal
        ("8B6914", "FBF0C8", "FEF7E4"),   # amber
    ]

    bdr = Border(**{s: Side(style="thin", color="DEE2E6")
                    for s in ["left", "right", "top", "bottom"]})

    wb = Workbook()
    ws = wb.active
    ws.title = "Forecast Submission"

    # Only include channels that have actual data in merged_data
    active_channels = [ch for ch in CHANNELS if ch in merged_data and merged_data[ch]]
    if not active_channels:
        active_channels = list(merged_data.keys())

    # Column layout: BASE_COLS_DRR + (months × channel) for each active channel
    n_base   = len(BASE_COLS_DRR)
    n_months = len(months)
    total_cols = n_base + n_months * len(active_channels)

    today_str = datetime.date.today().strftime("%d %b %Y")

    # ── Row 1: Title bar ──────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    tc = ws.cell(row=1, column=1,
                 value=f"Forecast — {qkey} — {username} — {today_str}  |  All Channels")
    tc.font      = Font(bold=True, color=C_WHITE, name="Calibri", size=12)
    tc.fill      = PatternFill(start_color=C_DARK, end_color=C_DARK, fill_type="solid")
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    # ── Row 2: Section headers (Product Info | Channel A | Channel B …) ───────
    # Product info block
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_base)
    h = ws.cell(row=2, column=1, value="PRODUCT INFO")
    h.font      = Font(bold=True, color=C_WHITE, name="Calibri", size=9)
    h.fill      = PatternFill(start_color=C_DARK, end_color=C_DARK, fill_type="solid")
    h.alignment = Alignment(horizontal="center", vertical="center")
    h.border    = bdr

    # One group per channel
    col_offset = n_base + 1
    for ci, ch in enumerate(active_channels):
        hdr_bg, _, _ = CHANNEL_PALETTES[ci % len(CHANNEL_PALETTES)]
        ws.merge_cells(start_row=2, start_column=col_offset,
                       end_row=2, end_column=col_offset + n_months - 1)
        cell = ws.cell(row=2, column=col_offset, value=f"FORECAST — {ch}")
        cell.font      = Font(bold=True, color=C_WHITE, name="Calibri", size=9)
        cell.fill      = PatternFill(start_color=hdr_bg, end_color=hdr_bg, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = bdr
        col_offset    += n_months
    ws.row_dimensions[2].height = 22

    # ── Row 3: Column sub-headers ─────────────────────────────────────────────
    for ci_b, h_name in enumerate(BASE_COLS_DRR, 1):
        cell = ws.cell(row=3, column=ci_b, value=h_name)
        cell.font      = Font(bold=True, color=C_WHITE, name="Calibri", size=9)
        cell.fill      = PatternFill(start_color=C_MID, end_color=C_MID, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = bdr

    col_offset = n_base + 1
    for ci, ch in enumerate(active_channels):
        _, sub_bg, _ = CHANNEL_PALETTES[ci % len(CHANNEL_PALETTES)]
        hdr_bg, _, _ = CHANNEL_PALETTES[ci % len(CHANNEL_PALETTES)]
        for m in months:
            cell = ws.cell(row=3, column=col_offset, value=m)
            cell.font      = Font(bold=True, color=C_WHITE, name="Calibri", size=9)
            cell.fill      = PatternFill(start_color=hdr_bg, end_color=hdr_bg, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = bdr
            col_offset    += 1
    ws.row_dimensions[3].height = 26

    # ── Rows 4+: Data ─────────────────────────────────────────────────────────
    for ri, prod in enumerate(drr_data, 4):
        even = ri % 2 == 0
        rid  = prod.get("_row_id", "")

        # Product info columns
        for ci_b, h_name in enumerate(BASE_COLS_DRR, 1):
            val  = prod.get(h_name, "")
            cell = ws.cell(row=ri, column=ci_b, value=val)
            cell.fill      = PatternFill(start_color=C_PALE if even else C_LIGHT,
                                         end_color=C_PALE if even else C_LIGHT,
                                         fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border    = bdr
            cell.font      = Font(name="Calibri", size=9)

        # Channel forecast columns
        col_offset = n_base + 1
        for ci, ch in enumerate(active_channels):
            _, data_bg_even, data_bg_odd = CHANNEL_PALETTES[ci % len(CHANNEL_PALETTES)]
            ch_data = merged_data.get(ch, {})
            row_vals = ch_data.get(rid, {})
            for m in months:
                v    = row_vals.get(m, "")
                cell = ws.cell(row=ri, column=col_offset, value=v if v != "" else "")
                cell.fill      = PatternFill(
                    start_color=data_bg_even if even else data_bg_odd,
                    end_color=data_bg_even if even else data_bg_odd,
                    fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = bdr
                cell.font      = Font(name="Calibri", size=9)
                if v != "":
                    cell.number_format = "#,##0.00"
                col_offset += 1

        ws.row_dimensions[ri].height = 18

    # Column widths
    for ci in range(1, n_base + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 16
    for ci in range(n_base + 1, total_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    ws.freeze_panes = get_column_letter(n_base + 1) + "4"
    wb.save(dest)


def _export_quarter_excel(qkey, q, dest):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    months   = QUARTERS[qkey]["months"]
    drr_data = q.get("drr_data") or []
    all_subs = _all_subs_for_quarter(qkey)
    all_members = [e for e in sorted(set(FORECAST_MEMBERS)) if e not in ADMINS]
    bdr = Border(**{s: Side(style="thin", color="DEE2E6") for s in ["left","right","top","bottom"]})
    wb  = Workbook()
    ws_sum = wb.active; ws_sum.title = "Summary"
    sum_headers = ["Channel / Member","Email","Status","Submitted At","Revision","Locked Months"]
    for ci, h in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bdr
    ws_sum.row_dimensions[1].height = 26
    ri = 2
    locked_months = _locked_months_in_quarter(qkey)
    for m_email in all_members:
        sub = all_subs.get(m_email, _sub_default())
        member_chs_exp = CHANNEL_MAP.get(m_email) or []
        ch  = ", ".join(member_chs_exp) if member_chs_exp else "Unassigned"
        row = [ch, sub.get("user_name", m_email), m_email,
               "Submitted" if sub.get("submitted") else "Pending",
               sub.get("submitted_at",""), sub.get("revision",0),
               ", ".join(locked_months)]
        for ci, v in enumerate(row[:len(sum_headers)], 1):
            cell = ws_sum.cell(row=ri, column=ci, value=v)
            cell.border = bdr
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ri += 1
    for ci, w in enumerate([18,28,32,14,20,8,18], 1):
        ws_sum.column_dimensions[get_column_letter(ci)].width = w
    ws = wb.create_sheet("All Submissions")
    prod_cols  = BASE_COLS_DRR
    drr_ch_cols = []
    for ch in CHANNELS:
        for dl in DRR_SHORT:
            drr_ch_cols.append(ch+" "+dl)
    meta_cols   = ["Member Name","Channel","Submitted At"]
    month_cols  = months
    all_headers = prod_cols + drr_ch_cols + meta_cols + month_cols
    col = 1
    def merged_header(ws, r, c, span, text, bg):
        if span > 1:
            ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+span-1)
        cell = ws.cell(row=r, column=c, value=text)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bdr
    merged_header(ws, 1, 1,               len(prod_cols),   "PRODUCT INFO",    "1B4332")
    merged_header(ws, 1, 1+len(prod_cols),len(drr_ch_cols), "HISTORICAL DRR",  "3730A3")
    merged_header(ws, 1, 1+len(prod_cols)+len(drr_ch_cols), len(meta_cols), "SUBMISSION", "2D6A4F")
    merged_header(ws, 1, 1+len(prod_cols)+len(drr_ch_cols)+len(meta_cols), len(month_cols), "FORECAST", "40916C")
    ws.row_dimensions[1].height = 24
    BG_PROD = "D8F3DC"; BG_DRR = "E0E7FF"; BG_META = "B7E4C7"; BG_MON = "DCFCE7"
    ch_color_map = {ch: bg for ch, bg in zip(CHANNELS,
        ["EEF2FF","DBEAFE","F3EEFF","FEF9C3","FDF2F8","FFF7ED","ECFEFF"])}
    for ci, h in enumerate(all_headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = Font(bold=True, name="Calibri", size=9,
                         color="1B4332" if ci <= len(prod_cols) else "374151")
        if ci <= len(prod_cols): bg = BG_PROD
        elif ci <= len(prod_cols)+len(drr_ch_cols):
            ch_idx = (ci - len(prod_cols) - 1) // len(DRR_SHORT)
            bg = ch_color_map.get(CHANNELS[ch_idx] if ch_idx < len(CHANNELS) else "", BG_DRR)
        elif ci <= len(prod_cols)+len(drr_ch_cols)+len(meta_cols): bg = BG_META
        else: bg = BG_MON
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bdr
    ws.row_dimensions[2].height = 32
    data_ri = 3
    for m_email in all_members:
        sub      = all_subs.get(m_email, _sub_default())
        raw_exp  = sub.get("data") or {}
        member_chs_ex = CHANNEL_MAP.get(m_email) or []
        ch_label_ex   = ", ".join(member_chs_ex) if member_chs_ex else "Unassigned"
        m_name        = sub.get("user_name", m_email.split("@")[0])
        submitted_at  = sub.get("submitted_at","")
        is_ch_keyed_ex = any(k in CHANNELS for k in raw_exp) if raw_exp else False
        export_channels = member_chs_ex if member_chs_ex else ["Unassigned"]
        for export_ch in export_channels:
            if is_ch_keyed_ex:
                saved_ch = raw_exp.get(export_ch, {})
            else:
                # Legacy flat data: show under first channel only
                saved_ch = raw_exp if export_ch == export_channels[0] else {}
            ch_display = export_ch
            even = data_ri % 2 == 0
            for prod in drr_data:
                rid     = prod.get("_row_id","")
                drr_all = prod.get("_drr", {})
                f_vals  = saved_ch.get(rid, {})
                row_vals = [prod.get(c,"") for c in prod_cols]
                for pch in CHANNELS:
                    ch_drr = drr_all.get(pch, {})
                    for dl in DRR_LABELS:
                        v = ch_drr.get(dl, "")
                        row_vals.append(round(float(v),2) if v else "")
                row_vals += [m_name, ch_display, submitted_at]
                for m in months:
                    row_vals.append(f_vals.get(m,""))
                for ci, val in enumerate(row_vals, 1):
                    cell = ws.cell(row=data_ri, column=ci, value=val)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = bdr
                    cell.font = Font(name="Calibri", size=9)
                    if ci <= len(prod_cols): bg = BG_PROD if even else "E8F5EC"
                    elif ci <= len(prod_cols)+len(drr_ch_cols):
                        ch_idx = (ci - len(prod_cols) - 1) // len(DRR_SHORT)
                        base   = ch_color_map.get(CHANNELS[ch_idx] if ch_idx < len(CHANNELS) else "", BG_DRR)
                        bg     = base if even else "F5F3FF"
                    elif ci <= len(prod_cols)+len(drr_ch_cols)+len(meta_cols):
                        bg = BG_META if even else "D8F3DC"
                    else:
                        bg = BG_MON if even else "F0FFF4"
                        if val == "" and sub.get("submitted"):
                            bg = "FEE2E2"
                    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                ws.row_dimensions[data_ri].height = 18
                data_ri += 1
    for ci in range(1, len(prod_cols)+1):
        ws.column_dimensions[get_column_letter(ci)].width = 16
    for ci in range(len(prod_cols)+1, len(prod_cols)+len(drr_ch_cols)+1):
        ws.column_dimensions[get_column_letter(ci)].width = 9
    for ci in range(len(prod_cols)+len(drr_ch_cols)+1, len(prod_cols)+len(drr_ch_cols)+len(meta_cols)+1):
        ws.column_dimensions[get_column_letter(ci)].width = 18
    for ci in range(len(prod_cols)+len(drr_ch_cols)+len(meta_cols)+1, len(all_headers)+1):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.freeze_panes = get_column_letter(len(prod_cols)+1)+"3"
    wb.save(dest)

# ── Insights dashboard ────────────────────────────────────────────────────────
@app.route("/insights")
@_require_admin
def insights():
    return render_template("insights.html",
        user_name=_name(),
        today=datetime.date.today().strftime("%A, %d %B %Y"),
        quarters=QUARTERS, channels=CHANNELS)

@app.route("/api/insights-data")
@_require_admin
def api_insights_data():
    from collections import defaultdict
    result = {}
    for qkey in QUARTERS:
        q = _q_get(qkey)
        if not q.get("initiated"): continue
        drr_data  = q.get("drr_data") or []
        all_subs  = _all_subs_for_quarter(qkey)
        members   = [e for e in FORECAST_MEMBERS if e not in ADMINS]
        months    = QUARTERS[qkey]["months"]
        submitted = [s for s in all_subs.values() if s.get("submitted")]
        pending   = len(members) - len(submitted)
        ch_drr_totals = defaultdict(lambda: defaultdict(float))
        cat_drr       = defaultdict(lambda: defaultdict(float))
        top_products  = []
        for row in drr_data:
            drr_all = row.get("_drr", {})
            cat     = row.get("Category", "Other")
            name    = row.get("Product Name", "")
            sku     = row.get("SKU", "")
            prod_total = 0
            prod_ch = {}
            for ch in CHANNELS:
                ch_drr = drr_all.get(ch, {})
                for lbl in DRR_LABELS:
                    v = ch_drr.get(lbl, 0) or 0
                    ch_drr_totals[ch][lbl] += v
                v7 = ch_drr.get("7 Days Overall DRR", 0) or 0
                cat_drr[cat][ch] += v7
                prod_total += v7
                prod_ch[ch] = round(v7, 1)
            top_products.append({"name":name,"sku":sku,"cat":cat,"total_7d":round(prod_total,1),"by_ch":prod_ch})
        top_products.sort(key=lambda x: x["total_7d"], reverse=True)
        ch_fill = {}
        for ch in CHANNELS:
            filled = 0; total_cells = 0
            for sub in all_subs.values():
                if sub.get("submitted"):
                    saved = sub.get("data") or {}
                    for rid in saved:
                        for m in months:
                            total_cells += 1
                            v = saved[rid].get(m, "")
                            if v != "" and v is not None: filled += 1
            ch_fill[ch] = {"filled":filled,"total":total_cells,
                           "pct":round(filled/total_cells*100,1) if total_cells else 0}
        result[qkey] = {
            "label":         QUARTERS[qkey]["label"],
            "sku_count":     len(drr_data),
            "total_members": len(members),
            "submitted":     len(submitted),
            "pending":       max(pending, 0),
            "ch_drr_totals": {ch: dict(v) for ch, v in ch_drr_totals.items()},
            "cat_drr":       {cat: dict(v) for cat, v in cat_drr.items()},
            "top_products":  top_products[:20],
            "ch_fill":       ch_fill,
        }
    return jsonify(result)

if __name__ == "__main__":
    print("\n  forecast.wbn — Wellbeing Nutrition v6")
    print("  http://localhost:5000\n")
    port = int(os.environ.get("PORT", 8000))
    app.run(host="0.0.0.0", port=port)