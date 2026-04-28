"""
excel_handler.py
Handles loading product data from Excel templates and saving submissions.
Supports both file paths and BytesIO buffers (for web/in-memory use).
Forecast Submission Tool — forecast.wbn | Wellbeing Nutrition
"""

import io
import os
import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── Schema ───────────────────────────────────────────────────────────────────

CHANNELS = ["D2C", "M-B2B", "M-B2C", "Retail", "Export", "Amazon", "Flipkart", "Myntra"]
DRR_PERIODS = [15, 30, 45, 60]

FIXED_COLUMNS = ["Category", "Sub Category", "Product Name / SKU"]

EDITABLE_COLUMNS = [
    f"{ch} {p}D"
    for ch in CHANNELS
    for p in DRR_PERIODS
]

ALL_COLUMNS = FIXED_COLUMNS + EDITABLE_COLUMNS


# ─── Sample Data ──────────────────────────────────────────────────────────────

def get_sample_products() -> list[dict]:
    return [
        {"Category": "Protein",   "Sub Category": "Whey Protein",     "Product Name / SKU": "WP-CHOC-1KG"},
        {"Category": "Protein",   "Sub Category": "Whey Protein",     "Product Name / SKU": "WP-VAN-1KG"},
        {"Category": "Protein",   "Sub Category": "Plant Protein",    "Product Name / SKU": "PP-CHOC-500G"},
        {"Category": "Protein",   "Sub Category": "Plant Protein",    "Product Name / SKU": "PP-BERRY-500G"},
        {"Category": "Vitamins",  "Sub Category": "Multivitamin",     "Product Name / SKU": "MV-DAILY-60TAB"},
        {"Category": "Vitamins",  "Sub Category": "Vitamin D3",       "Product Name / SKU": "VD3-2000IU-90TAB"},
        {"Category": "Vitamins",  "Sub Category": "Vitamin C",        "Product Name / SKU": "VC-1000MG-60TAB"},
        {"Category": "Wellness",  "Sub Category": "Omega 3",          "Product Name / SKU": "OM3-1000MG-60CAP"},
        {"Category": "Wellness",  "Sub Category": "Probiotics",       "Product Name / SKU": "PRO-10B-30CAP"},
        {"Category": "Wellness",  "Sub Category": "Ashwagandha",      "Product Name / SKU": "ASH-600MG-60CAP"},
        {"Category": "Nutrition", "Sub Category": "Meal Replacement", "Product Name / SKU": "MR-CHOC-500G"},
        {"Category": "Nutrition", "Sub Category": "Meal Replacement", "Product Name / SKU": "MR-VAN-500G"},
        {"Category": "Nutrition", "Sub Category": "Energy Bar",       "Product Name / SKU": "EB-CHOC-12PACK"},
        {"Category": "Nutrition", "Sub Category": "Protein Bar",      "Product Name / SKU": "PB-PNUT-12PACK"},
        {"Category": "Sports",    "Sub Category": "Pre-Workout",      "Product Name / SKU": "PW-CITRUS-300G"},
        {"Category": "Sports",    "Sub Category": "BCAA",             "Product Name / SKU": "BCAA-WATERMELON-300G"},
        {"Category": "Sports",    "Sub Category": "Creatine",         "Product Name / SKU": "CR-MONO-300G"},
        {"Category": "Beauty",    "Sub Category": "Collagen",         "Product Name / SKU": "COL-MARINE-200G"},
        {"Category": "Beauty",    "Sub Category": "Biotin",           "Product Name / SKU": "BIO-5000MCG-60TAB"},
        {"Category": "Kids",      "Sub Category": "Kids Vitamins",    "Product Name / SKU": "KV-GUMMY-60PC"},
    ]


# ─── Template Creation ────────────────────────────────────────────────────────

def create_template_excel(dest) -> None:
    """
    Create a blank Excel template.
    dest: file path (str) or BytesIO buffer.
    """
    products = get_sample_products()

    wb = Workbook()
    ws = wb.active
    ws.title = "Forecast Template"

    header_fill = PatternFill(start_color="2D6A4F", end_color="2D6A4F", fill_type="solid")
    fixed_fill  = PatternFill(start_color="D8F3DC", end_color="D8F3DC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for col_idx, col_name in enumerate(ALL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border

    for row_idx, product in enumerate(products, 2):
        for col_idx, col_name in enumerate(ALL_COLUMNS, 1):
            val  = product.get(col_name, 0.0)
            cell = ws.cell(row=row_idx, column=col_idx, value=val if col_name in FIXED_COLUMNS else 0.0)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border
            if col_name in FIXED_COLUMNS:
                cell.fill = fixed_fill

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 24
    for i in range(4, len(ALL_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 12

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "D2"

    wb.save(dest)


# ─── Load from Excel ──────────────────────────────────────────────────────────

def load_products_from_excel(source) -> list[dict]:
    """
    Load product rows from an Excel file.
    source: file path (str) or BytesIO buffer.
    """
    try:
        df = pd.read_excel(source, sheet_name=0)
        df.columns = [str(c).strip() for c in df.columns]

        for fc in FIXED_COLUMNS:
            if fc not in df.columns:
                raise ValueError(
                    f"Missing required column: '{fc}'.\n"
                    "Use 'Download Template' to get the correct format."
                )

        for ec in EDITABLE_COLUMNS:
            if ec not in df.columns:
                df[ec] = 0.0

        df = df[ALL_COLUMNS].copy()
        for ec in EDITABLE_COLUMNS:
            df[ec] = pd.to_numeric(df[ec], errors="coerce").fillna(0.0)

        return df.to_dict("records")

    except Exception as e:
        raise RuntimeError(f"Failed to load Excel file:\n{str(e)}")


# ─── Save Submission ──────────────────────────────────────────────────────────

def save_submission_excel(data: list[dict], username: str, dest) -> None:
    """
    Save filled forecast data as a formatted Excel file.
    dest: file path (str) or BytesIO buffer.
    """
    today     = datetime.date.today()
    dark_green  = "1B4332"
    mid_green   = "2D6A4F"
    pale_green  = "D8F3DC"
    accent_green = "B7E4C7"
    white       = "FFFFFF"
    light_grey  = "F8F9FA"
    border_grey = "DEE2E6"

    thin_border = Border(
        left=Side(style="thin", color=border_grey),
        right=Side(style="thin", color=border_grey),
        top=Side(style="thin", color=border_grey),
        bottom=Side(style="thin", color=border_grey),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Forecast Submission"

    # Row 1: Title
    ws.merge_cells(f"A1:{get_column_letter(len(ALL_COLUMNS))}1")
    cell = ws["A1"]
    cell.value     = f"Forecast Submission — {today.strftime('%d %B %Y')} — {username}  |  Wellbeing Nutrition"
    cell.font      = Font(bold=True, color=white, name="Calibri", size=13)
    cell.fill      = PatternFill(start_color=dark_green, end_color=dark_green, fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[1].height = 30

    # Row 2: Channel group headers
    for i in range(1, len(FIXED_COLUMNS) + 1):
        ws.cell(row=2, column=i).fill = PatternFill(start_color=mid_green, end_color=mid_green, fill_type="solid")

    col_offset = len(FIXED_COLUMNS) + 1
    for ch in CHANNELS:
        end_col = col_offset + len(DRR_PERIODS) - 1
        ws.merge_cells(start_row=2, start_column=col_offset, end_row=2, end_column=end_col)
        cell           = ws.cell(row=2, column=col_offset, value=ch)
        cell.font      = Font(bold=True, color=white, name="Calibri", size=10)
        cell.fill      = PatternFill(start_color=mid_green, end_color=mid_green, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col_offset    += len(DRR_PERIODS)
    ws.row_dimensions[2].height = 24

    # Row 3: Column headers
    for col_idx, col_name in enumerate(ALL_COLUMNS, 1):
        cell           = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font      = Font(bold=True, name="Calibri", size=9, color=dark_green)
        cell.fill      = PatternFill(
            start_color=pale_green if col_name in FIXED_COLUMNS else accent_green,
            end_color=pale_green   if col_name in FIXED_COLUMNS else accent_green,
            fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border
    ws.row_dimensions[3].height = 36

    # Data rows
    for row_idx, row_data in enumerate(data, 4):
        even = row_idx % 2 == 0
        for col_idx, col_name in enumerate(ALL_COLUMNS, 1):
            val  = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border
            if col_name in FIXED_COLUMNS:
                cell.fill = PatternFill(start_color=pale_green, end_color=pale_green, fill_type="solid")
                cell.font = Font(name="Calibri", size=9, bold=(col_name == "Product Name / SKU"))
            else:
                bg = white if even else light_grey
                cell.fill          = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                cell.font          = Font(name="Calibri", size=9)
                cell.number_format = "#,##0.00"
        ws.row_dimensions[row_idx].height = 20

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 24
    for i in range(4, len(ALL_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 11

    ws.freeze_panes = "D4"
    wb.save(dest)


# ─── Audit Log ────────────────────────────────────────────────────────────────

def save_submission_log(username: str, filename: str, row_count: int, log_dir: str = ".") -> None:
    """Append a submission record to a local CSV audit log."""
    log_path = os.path.join(log_dir, "forecast_submission_log.csv")
    now      = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    entry    = pd.DataFrame([{
        "Timestamp":      now,
        "Username":       username,
        "Rows Submitted": row_count,
        "File":           filename,
    }])
    if os.path.exists(log_path):
        combined = pd.concat([pd.read_csv(log_path), entry], ignore_index=True)
    else:
        combined = entry
    combined.to_csv(log_path, index=False)
